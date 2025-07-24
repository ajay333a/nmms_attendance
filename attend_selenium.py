import io
import re
import time
from urllib.parse import urljoin
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

# Constants
BASE_URL = "https://mnregaweb4.nic.in/nregaarch/View_NMMS_atten_date_new.aspx?fin_year=2024-2025&Digest=HNrisV4bhHnb7Gve3mAKYQ"
STATE_VALUE = '15'  # Karnataka
DISTRICT_NAME = 'BALLARI'
BLOCK_NAME = 'SIRUGUPPA'
TALUK_NAME = 'Siruguppa'
DISTRICT_LABEL = 'Ballari'

def get_table_by_id_or_div(soup, table_id='grdTable', div_id='RepPr1'):
    table = soup.find('table', {'id': table_id})
    if not table:
        div = soup.find('div', {'id': div_id})
        if div:
            table = div.find('table')
    return table

def get_link_from_table(table, match_col_idx, match_text):
    for row in table.find_all('tr'):
        cols = row.find_all('td')
        if len(cols) > match_col_idx and cols[match_col_idx].get_text(strip=True).upper() == match_text.upper():
            a = cols[match_col_idx].find('a', href=True)
            if a:
                return a['href']
    return None

def get_panchayath_link(table, panchayath_name):
    for row in table.find_all('tr'):
        cols = row.find_all('td')
        if len(cols) >= 4:
            s_no = cols[0].get_text(strip=True)
            if not s_no.isdigit():
                continue
            panch_name = cols[1].get_text(strip=True).upper()
            muster_rolls_a = cols[3].find('a', href=True)
            href = muster_rolls_a['href'] if muster_rolls_a else None
            if panch_name == panchayath_name and href:
                return href
    return None

def get_muster_roll_rows(muster_table, choice, workcodes=None, workcode_idx=None, muster_no_idx=None):
    rows_to_save = []
    if workcodes is None:
        workcodes = []
    for row in muster_table.find_all('tr')[1:]:
        cols = row.find_all('td')
        if len(cols) > muster_no_idx:
            if choice == 'all':
                muster_a = cols[muster_no_idx].find('a', href=True)
                if muster_a:
                    muster_href = muster_a['href']
                    rows_to_save.append((cols, muster_href))
            elif choice == 'work' and workcodes and workcode_idx is not None:
                current_workcode = cols[workcode_idx].get_text(strip=True)
                if len(cols) > workcode_idx and any(wc in current_workcode for wc in workcodes):
                    muster_a = cols[muster_no_idx].find('a', href=True)
                    if muster_a:
                        muster_href = muster_a['href']
                        rows_to_save.append((cols, muster_href))
    return rows_to_save

def save_attendance_excel(wb, ws, img_wb, img_ws, panchayath_name, attendance_date):
    wb.save(f"muster_rolls_{panchayath_name}_{attendance_date.replace('/', '_')}.xlsx")
    img_wb.save(f"muster_roll_images_{panchayath_name}_{attendance_date.replace('/', '_')}.xlsx")
    print(f"Saved muster_rolls_{panchayath_name}_{attendance_date.replace('/', '_')}.xlsx")
    print(f"Saved muster_roll_images_{panchayath_name}_{attendance_date.replace('/', '_')}.xlsx")

def save_raw_excel(rows_to_save, panchayath_name, attendance_date, muster_no_idx, workcode_idx, panchayath_url, muster_data_cache):
    raw_wb = openpyxl.Workbook()
    raw_ws = raw_wb.active
    raw_ws.append([
        "Taluk", "Panchayath", "Work Code", "Muster Roll No", "Job Card No", "Worker Name", "Gender", "Attendance", "Attendance Date"
    ])
    for cols, muster_href in rows_to_save:
        muster_url = urljoin(panchayath_url, muster_href)
        attendance_data, _, _, _ = muster_data_cache.get(muster_url, (None, None, None, None))
        muster_roll_no = cols[muster_no_idx].get_text(strip=True)
        work_code = cols[workcode_idx].get_text(strip=True)
        for att_row in attendance_data or []:
            worker_name_full = att_row[2] if len(att_row) > 2 else ''
            if worker_name_full.endswith(')') and '(' in worker_name_full:
                name_part = worker_name_full[:worker_name_full.rfind('(')].strip()
                gender_part = worker_name_full[worker_name_full.rfind('(')+1:-1].strip()
            else:
                name_part = worker_name_full
                gender_part = ''
            raw_ws.append([
                TALUK_NAME,
                panchayath_name,
                work_code,
                muster_roll_no,
                att_row[1] if len(att_row) > 1 else '',
                name_part,
                gender_part,
                att_row[4] if len(att_row) > 4 else '',
                att_row[3] if len(att_row) > 3 else '',
            ])
    return raw_wb

def find_col_idx(header_cols, search):
    search_clean = re.sub(r'[^a-zA-Z0-9]', '', search.lower())
    for i, h in enumerate(header_cols):
        h_clean = re.sub(r'[^a-zA-Z0-9]', '', h.lower())
        if search_clean in h_clean:
            return i
    return None

def get_attendance_data(driver, url):
    driver.get(url)
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    work_name = None
    for b in soup.find_all('b'):
        if b.text.strip().startswith('Work Name'):
            next_text = b.next_sibling
            if next_text:
                work_name = str(next_text).strip(' :\u00a0-')
            break
    if not work_name:
        work_name_elem = soup.find(id="ContentPlaceHolder1_lbl_dtl")
        if work_name_elem:
            work_name = work_name_elem.text.strip()
    attendance_data = []
    tables = soup.find_all('table')
    if not tables:
        print("No tables found on the page.")
        return None, None, work_name, None
    attendance_table = tables[-1]
    rows = attendance_table.find_all('tr')
    header_cells = [th.text.strip() for th in rows[0].find_all(['th', 'td'])]
    col_map = {name: idx for idx, name in enumerate(header_cells)}
    for row in rows[1:]:
        cols = row.find_all('td')
        if cols and any(c.get_text(strip=True) for c in cols):
            name_td = ''
            for td in cols:
                span = td.find('span', id=lambda x: x and 'lbl_workerName_' in x)
                if span:
                    name_td = span.get_text(strip=True)
                    break
            extracted = [
                cols[col_map.get('S.No', -1)].get_text(strip=True) if 'S.No' in col_map else '',
                cols[col_map.get('Job Card No', -1)].get_text(strip=True) if 'Job Card No' in col_map else '',
                name_td,
                cols[col_map.get('Attendance Date', -1)].get_text(strip=True) if 'Attendance Date' in col_map else '',
                cols[col_map.get('Present/Absent', -1)].get_text(strip=True) if 'Present/Absent' in col_map else ''
            ]
            attendance_data.append(extracted)
    photo_url = None
    img_link = soup.find('a', text='Click here for large image')
    if img_link and img_link.has_attr('href'):
        photo_url = urljoin(driver.current_url, img_link['href'])
    return attendance_data, photo_url, work_name, header_cells

def download_photo(driver, url):
    if not url:
        return None
    try:
        driver.get(url)
        # This assumes the image is the only thing on the page
        img_bytes = io.BytesIO(driver.find_element(By.TAG_NAME, 'img').screenshot_as_png)
        return img_bytes
    except Exception as e:
        print(f"Error downloading photo: {e}")
        return None

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from selenium.common.exceptions import TimeoutException

from selenium.common.exceptions import TimeoutException, StaleElementReferenceException

def resilient_click(driver, by, value, retries=3, delay=2):
    """
    A wrapper function that attempts to find and click an element,
    retrying if a StaleElementReferenceException is caught.
    """
    for i in range(retries):
        try:
            element = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((by, value))
            )
            element.click()
            return
        except StaleElementReferenceException:
            print(f"Stale element reference caught. Retrying ({i+1}/{retries})...")
            time.sleep(delay)
        except TimeoutException:
            raise TimeoutException(f"Element with {by}='{value}' not found or clickable after waiting.")
    raise StaleElementReferenceException(f"Element with {by}='{value}' was stale after {retries} retries.")

def get_work_codes(driver, attendance_date, panchayath_name):
    driver.get(BASE_URL)
    wait = WebDriverWait(driver, 40)

    # --- Step 1: Initial Page ---
    Select(wait.until(EC.presence_of_element_located((By.NAME, 'ctl00$ContentPlaceHolder1$ddlstate')))).select_by_value(STATE_VALUE)
    wait.until(EC.presence_of_element_located((By.XPATH, f"//select[@name='ctl00$ContentPlaceHolder1$ddl_attendance']/option[normalize-space()='{attendance_date}']")))
    Select(driver.find_element(By.NAME, 'ctl00$ContentPlaceHolder1$ddl_attendance')).select_by_value(attendance_date)
    resilient_click(driver, By.NAME, 'ctl00$ContentPlaceHolder1$btn_showreport')

    # --- Step 2-4: Navigate Hierarchy ---
    resilient_click(driver, By.LINK_TEXT, 'KARNATAKA')
    wait.until(EC.presence_of_element_located((By.LINK_TEXT, DISTRICT_NAME)))
    
    resilient_click(driver, By.LINK_TEXT, DISTRICT_NAME)
    wait.until(EC.presence_of_element_located((By.LINK_TEXT, BLOCK_NAME)))

    resilient_click(driver, By.LINK_TEXT, BLOCK_NAME)
    wait.until(EC.presence_of_element_located((By.ID, 'RepPr1')))

    # --- Step 5: Panchayath Page ---
    try:
        panchayath_xpath = f"//table//tr[td[2][normalize-space(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'))='{panchayath_name.lower()}']]/td[4]/a"
        panchayath_link = wait.until(EC.presence_of_element_located((By.XPATH, panchayath_xpath)))
        panchayath_url = panchayath_link.get_attribute('href')
        resilient_click(driver, By.XPATH, panchayath_xpath)
    except TimeoutException:
        raise Exception("The selected Panchayath has not generated any Muster Roll for the chosen date.")

    # --- Step 6: Muster Roll Page ---
    wait.until(EC.presence_of_element_located((By.ID, 'RepPr1')))
    page_source = driver.page_source
    soup = BeautifulSoup(page_source, 'html.parser')
    muster_table = soup.find('div', {'id': 'RepPr1'}).find('table')
    if not muster_table:
        raise Exception("Could not find muster roll table.")
    
    header_row = muster_table.find('tr')
    header_cols = [th.get_text(strip=True).replace('\u00a0', ' ').strip().lower() for th in header_row.find_all(['th', 'td'])]
    workcode_idx = find_col_idx(header_cols, 'work code')
    muster_no_idx = find_col_idx(header_cols, 'mustroll no')
    if workcode_idx is None or muster_no_idx is None:
        raise Exception("Could not find required columns in muster roll table header.")

    all_workcodes = set()
    for row in muster_table.find_all('tr')[1:]:
        cols = row.find_all('td')
        if len(cols) > workcode_idx and len(cols) > muster_no_idx and cols[muster_no_idx].find('a', href=True):
            all_workcodes.add(cols[workcode_idx].get_text(strip=True))
            
    return sorted(list(all_workcodes)), page_source, panchayath_url, workcode_idx, muster_no_idx




def run_scraper(driver, page_source, panchayath_url, panchayath_name, attendance_date, workcode_idx, muster_no_idx, selected_work_codes, status_callback):
    soup = BeautifulSoup(page_source, 'html.parser')
    muster_table = soup.find('div', {'id': 'RepPr1'}).find('table')
    
    choice = 'all' if 'all' in selected_work_codes else 'work'
    rows_to_save = get_muster_roll_rows(muster_table, choice, selected_work_codes, workcode_idx, muster_no_idx)
    
    if not rows_to_save:
        raise Exception("No muster roll data found for the selection.")

    # Excel setup
    wb = openpyxl.Workbook()
    ws = wb.active
    row_cursor = 1
    ws.cell(row=row_cursor, column=1, value="District:").font = Font(bold=True)
    ws.cell(row=row_cursor, column=2, value=DISTRICT_LABEL)
    ws.cell(row=row_cursor, column=3, value="Taluk/Block:").font = Font(bold=True)
    ws.cell(row=row_cursor, column=4, value=TALUK_NAME)
    row_cursor += 1
    ws.cell(row=row_cursor, column=1, value="Panchayath:").font = Font(bold=True)
    ws.cell(row=row_cursor, column=2, value=panchayath_name)
    row_cursor += 1
    workcode_row_idx_att = row_cursor
    ws.cell(row=workcode_row_idx_att, column=1, value="Work code:").font = Font(bold=True)
    ws.cell(row=workcode_row_idx_att, column=3, value="Work Name:").font = Font(bold=True)
    row_cursor += 1
    attendance_header_written = False
    first_muster_processed = False

    # Image-only Excel setup
    img_wb = openpyxl.Workbook()
    img_ws = img_wb.active
    img_row_cursor = 1
    img_bytes_refs = []
    img_ws.cell(row=img_row_cursor, column=1, value="District:").font = Font(bold=True)
    img_ws.cell(row=img_row_cursor, column=2, value=DISTRICT_LABEL)
    img_ws.cell(row=img_row_cursor, column=3, value="Taluk/Block:").font = Font(bold=True)
    img_ws.cell(row=img_row_cursor, column=4, value=TALUK_NAME)
    img_row_cursor += 1
    img_ws.cell(row=img_row_cursor, column=1, value="Panchayath:").font = Font(bold=True)
    img_ws.cell(row=img_row_cursor, column=2, value=panchayath_name)
    img_row_cursor += 1
    workcode_row_idx = img_row_cursor
    img_ws.cell(row=workcode_row_idx, column=1, value="Work code:").font = Font(bold=True)
    img_ws.cell(row=workcode_row_idx, column=3, value="Work Name:").font = Font(bold=True)
    img_row_cursor += 1
    header_row_idx = img_row_cursor
    img_ws.cell(row=header_row_idx, column=1, value='Muster Roll No').font = Font(bold=True)
    img_ws.cell(row=header_row_idx, column=2, value='Image').font = Font(bold=True)
    img_row_cursor += 1

    muster_data_cache = {}
    total_rows = len(rows_to_save)
    for i, (cols, muster_href) in enumerate(rows_to_save):
        muster_url = urljoin(panchayath_url, muster_href)
        
        status_callback(f"Processing muster roll {i+1}/{total_rows}...", (i+1)/total_rows * 100)
        
        attendance_data, photo_url, work_name, header_cells = get_attendance_data(driver, muster_url)
        muster_data_cache[muster_url] = (attendance_data, photo_url, work_name, header_cells)
        
        img_bytes = download_photo(driver, photo_url) if photo_url else None
        muster_roll_no = cols[muster_no_idx].get_text(strip=True)

        if not attendance_header_written and header_cells:
            if not first_muster_processed:
                first_work_code = cols[workcode_idx].get_text(strip=True)
                first_work_name = work_name or ''
                ws.cell(row=workcode_row_idx_att, column=2, value=first_work_code)
                ws.cell(row=workcode_row_idx_att, column=4, value=first_work_name)
                img_ws.cell(row=workcode_row_idx, column=2, value=first_work_code)
                img_ws.cell(row=workcode_row_idx, column=4, value=first_work_name)
                first_muster_processed = True
            ws.cell(row=row_cursor, column=1, value="Muster Roll No").font = Font(bold=True)
            for col_idx, header in enumerate(header_cells, 2):
                ws.cell(row=row_cursor, column=col_idx, value=header).font = Font(bold=True)
            row_cursor += 1
            attendance_header_written = True
        
        if attendance_data:
            for att_row in attendance_data:
                ws.cell(row=row_cursor, column=1, value=muster_roll_no)
                for col_idx, val in enumerate(att_row, 2):
                    ws.cell(row=row_cursor, column=col_idx, value=val)
                row_cursor += 1
        
        if img_bytes:
            img_bytes.seek(0)
            img = XLImage(img_bytes)
            img_cell = f"H{row_cursor-len(attendance_data) if attendance_data else row_cursor}"
            ws.add_image(img, img_cell)
            row_cursor += 3
        else:
            row_cursor += 2
        row_cursor += 2

        start_img_row = img_row_cursor
        img_ws.cell(row=img_row_cursor, column=1, value=muster_roll_no).font = Font(bold=True, size=18)
        if img_bytes:
            img_bytes.seek(0)
            img_bytes_for_imgwb = io.BytesIO(img_bytes.getbuffer())
            img2 = XLImage(img_bytes_for_imgwb)
            img_ws.add_image(img2, f"B{img_row_cursor}")
            img_bytes_refs.append(img_bytes_for_imgwb)
            img_height_rows = 20
            end_img_row = img_row_cursor + img_height_rows - 1
            img_ws.merge_cells(start_row=start_img_row, start_column=1, end_row=end_img_row, end_column=1)
            img_ws.cell(row=start_img_row, column=1).alignment = Alignment(vertical='center', horizontal='center')
            img_row_cursor += img_height_rows
        else:
            end_img_row = img_row_cursor
            img_row_cursor += 3
            img_ws.merge_cells(start_row=start_img_row, start_column=1, end_row=end_img_row, end_column=1)
            img_ws.cell(row=start_img_row, column=1).alignment = Alignment(vertical='center', horizontal='center')
        img_row_cursor += 2

    raw_wb = save_raw_excel(rows_to_save, panchayath_name, attendance_date, muster_no_idx, workcode_idx, panchayath_url, muster_data_cache)

    wb_io = io.BytesIO()
    wb.save(wb_io)
    wb_io.seek(0)

    img_wb_io = io.BytesIO()
    img_wb.save(img_wb_io)
    img_wb_io.seek(0)
    
    raw_wb_io = io.BytesIO()
    raw_wb.save(raw_wb_io)
    raw_wb_io.seek(0)

    return wb_io, img_wb_io, raw_wb_io

def main():
    # This function is for standalone script execution
    pass

if __name__ == "__main__":
    main()
